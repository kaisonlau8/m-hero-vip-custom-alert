"""解析 DMS 导出的保养提醒任务 Excel。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

SHEET_NAME = "客户回访任务中心"
REQUIRED_COLUMNS = ("VIN", "任务编码", "任务类型", "创建日期")

# Excel 列名 → 任务 dict 键
OPTIONAL_COLUMNS: dict[str, str] = {
    "区域": "region",
    "门店编码": "store_code",
    "门店名称": "store_name",
    "任务状态": "task_status",
    "售后车系": "aftersales_series",
    "用车人名称": "driver_name",
    "用车人电话": "driver_phone",
    "车主名称": "owner_name",
    "车主电话": "owner_phone",
    "下次预约时间": "next_appointment_at",
    "预约单号": "appointment_no",
    "到期日期": "due_at",
    "首次回访日期": "first_followup_at",
    "首次回访人": "first_followup_by",
    "状态": "dms_valid_status",
    "关闭时间": "closed_at",
}


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_datetime_str(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        # 日期无时分秒时仍输出完整可读串
        try:
            if getattr(value, "hour", None) == 0 and getattr(value, "minute", None) == 0 and getattr(value, "second", None) == 0:
                # 仍保留时间，与导出一致
                return value.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return _cell_str(value)


def import_maintenance_reminder_xlsx(xlsx_path: str | Path) -> list[dict]:
    path = Path(xlsx_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Excel 不存在: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]

    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [_cell_str(h) for h in header_row]
    index = {name: i for i, name in enumerate(headers) if name}
    missing = [c for c in REQUIRED_COLUMNS if c not in index]
    if missing:
        wb.close()
        raise RuntimeError(f"Excel 缺少列 {missing}，实际表头: {headers}")

    tasks: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not row:
            continue
        vin = _cell_str(row[index["VIN"]]).upper()
        task_code = _cell_str(row[index["任务编码"]])
        if not vin or not task_code:
            continue

        task: dict[str, Any] = {
            "vin": vin,
            "task_code": task_code,
            "task_type": _cell_str(row[index["任务类型"]]),
            "created_at": _cell_datetime_str(row[index["创建日期"]]),
        }
        for col_name, key in OPTIONAL_COLUMNS.items():
            if col_name not in index:
                task[key] = ""
                continue
            raw = row[index[col_name]]
            if col_name in (
                "下次预约时间",
                "到期日期",
                "首次回访日期",
                "关闭时间",
            ):
                task[key] = _cell_datetime_str(raw)
            else:
                task[key] = _cell_str(raw)
        tasks.append(task)

    wb.close()
    return tasks


def find_latest_download(download_dir: str | Path) -> Path | None:
    directory = Path(download_dir)
    if not directory.exists():
        return None
    files = sorted(
        [
            p
            for p in directory.glob("*.xlsx")
            if p.is_file() and not p.name.startswith("~$")
        ],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None


if __name__ == "__main__":
    import json
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        raise SystemExit("用法: python import_excel.py <xlsx>")
    data = import_maintenance_reminder_xlsx(path)
    print(json.dumps({"count": len(data), "sample": data[:3]}, ensure_ascii=False, indent=2))
